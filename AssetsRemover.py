import openpyxl
import os

##################################################################################################
##############################使用前请替换「unity工程」根目录路径######################################
##################################################################################################
#unity 工程项目根目录
avatar_root_path = r'Z:\Users\ryanco\Projects\AndoidProject\wonder_party\avatarProject'
##################################################################################################
##################################################################################################
##################################################################################################


excel_path = avatar_root_path + r'\Assets/Editor\AssetsManagement\SuitExcel\服饰元表.xlsx'
project_path = avatar_root_path + r'\\'
prefab_path = avatar_root_path + r'\Assets\Art\BundleResources\Dress'
assets_path = avatar_root_path + r'\Assets\Art\model\coat'
sprite_path = avatar_root_path + r'\Assets\Art\BundleResources\Sprites'


wb = openpyxl.load_workbook(excel_path)
nws = wb['New_Added_Assets_List']

def GetLastRow(ws):
    for row in range(1, ws.max_row + 100000):
        if (ws.cell(row, 3).value is None): #check time stamp for last row
            last_row = row;  # this last_row value is actual plus one by actual last row in excel, cus py access col by minis one
            print("last row is : " + str(last_row))
            return last_row

def CheckAssetToRemove(nws):
    #loop whole excel
    for row in range(2, last_row):  # start at 2 , cus first row is not the actual info
        file_col_list = [7, 10, 12, 14, 16]
        print(nws.cell(row, 4).value)
        if nws.cell(row, 4).value == None:
            print("Empty : " + str(row))
            for col in file_col_list:
                if not nws.cell(row, col).value == "0" and not str(nws.cell(row, col).value).__contains__("旧资源"):
                    #detele file in unity
                    DeleteCell(nws.cell(row, col).value)

            #detele whole row
            nws.delete_rows(row)
            wb.save(excel_path)

def DeleteCell(cell_value):
    cell_path_list = str(cell_value).split('|-|')
    for path in cell_path_list:
        if len(path) < 5:
            # 绕过序号
            continue

        #delete every file
        delete_file_path = project_path + path.replace('/', '\\')
        os.remove(delete_file_path)
        print("Delete : " + delete_file_path)





last_row = GetLastRow(nws)
CheckAssetToRemove(nws)