import openpyxl
import os
from pathlib import Path
from openpyxl.styles import Color, PatternFill, Font, Border

redFill = PatternFill(start_color='FF5645',end_color='FF5645',fill_type='solid')
blueFill = PatternFill(start_color='2FA0FF',end_color='2FA0FF',fill_type='solid')

excel_path = r'/Users/ryanco/Desktop/资源元表/服饰元表Excel.xlsx'
project_path = r'/Users/ryanco/Projects/AndoidProject/wonder_party/avatarProject/'
prefab_path = r'/Users/ryanco/Projects/AndoidProject/wonder_party/avatarProject/Assets/Art/BundleResources/Dress'
assets_path = r'/Users/ryanco/Projects/AndoidProject/wonder_party/avatarProject/Assets/Art/model/coat'
sprite_path = r'/Users/ryanco/Projects/AndoidProject/wonder_party/avatarProject/Assets/Art/BundleResources/Sprites'

#get excel
wb = openpyxl.load_workbook(excel_path)
ws = wb['AssetsInfo - Assets_Art_model_c']

def WriteAssetInfo():
    # r=root, d=directories, f = files
    count = 0;
    for r, d, f in os.walk(prefab_path):
        f.sort()
        for file in f:
            if file.endswith(".prefab"):

                ### NameCode
                namecode = Path(os.path.join(r, file)).stem
                ws.cell(2 + count, 2).value = namecode

                ### AssetType
                dress_type = os.path.join(r, file).split('/')[-2]
                ws.cell(2 + count, 3).value = dress_type



                ### Prafab Path
                ws.cell(2 + count, 5).value = os.path.join(r, file).replace(project_path, "")

                ### Assets Folder Path
                ws.cell(2 + count, 5).value = ""


                ## Texture Path
                FindAssets(count, namecode, "png", 8, "贴图")

                ## Mesh Path
                FindAssets(count, namecode, "fbx", 10, "模型")

                ## Mat Path
                FindAssets(count, namecode, "mat", 12, "材质")

                ## Sprite Path
                FindSprite(count, namecode, "png", 14, "缩略图")

                ### SubType
                sprite_value = ws.cell(2 + count, 14).value
                AssetSubType(count, namecode, dress_type, sprite_value)




                count += 1
    wb.save("/Users/ryanco/Desktop/资源元表/服饰元表Excel.xlsx")

def FindAssets(count, namecode, asset_type, col, lost_text):
    asset_list = []

    for r, d, f in os.walk(assets_path):
        for file in f:
            if file.__contains__(namecode) and file.lower().endswith("." + asset_type):
                asset_list.append(os.path.join(r, file).replace(project_path, ""))

    if len(asset_list) == 0:
        ws.cell(2 + count, col).value = lost_text + "缺失"
        ws.cell(2 + count, col).fill = redFill
    else:
        if len(asset_list) > 1:
            # print()
            ws.cell(2 + count, col).fill = blueFill
        ws.cell(2 + count, col).value = str(len(asset_list))
        for path in asset_list:
            ws.cell(2 + count, col).value += "|-|"
            ws.cell(2 + count, col).value += path

def FindSprite(count, namecode, asset_type, col, lost_text):
    asset_list = []

    for r, d, f in os.walk(sprite_path):
        for file in f:
            if file.__contains__(namecode) and file.lower().endswith("." + asset_type):
                asset_list.append(os.path.join(r, file).replace(project_path, ""))

    if len(asset_list) == 0:
        ws.cell(2 + count, col).value = lost_text + "缺失"
        ws.cell(2 + count, col).fill = redFill
    else:
        if len(asset_list) > 1:
            # print()
            ws.cell(2 + count, col).fill = blueFill
        ws.cell(2 + count, col).value = str(len(asset_list))
        for path in asset_list:
            ws.cell(2 + count, col).value += "|-|"
            ws.cell(2 + count, col).value += path

def AssetSubType(count, namecode, dress_type, sprite_value):
    if sprite_value == "缩略图缺失":
        return

    if dress_type == "headwear" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通头饰"
        return
    elif dress_type == "headwear" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图头饰"
        return

    if dress_type == "baldric" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通背包"
        return
    elif dress_type == "baldric" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图背包"
        return

    if dress_type == "glasses" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通眼镜"
        return
    elif dress_type == "glasses" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图眼镜"
        return

    if dress_type == "pants" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通裤子"
        return
    elif dress_type == "pants" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图裤子"
        return

    if dress_type == "suit" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通套装"
        return
    elif dress_type == "suit" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图套装"
        return

    if dress_type == "shoes" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通鞋子"
        return
    elif dress_type == "shoes" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图鞋子"
        return

    if dress_type == "hair" and namecode[0].lower() == "m":
        ws.cell(2 + count, 4).value = "帽子头发"
        return
    elif dress_type == "hair" and namecode[-1].lower() == "a":
        ws.cell(2 + count, 4).value = "旧版本头发"
        return
    elif dress_type == "hair" and namecode[-1].lower() == "s":
        ws.cell(2 + count, 4).value = "旧版本头发"
        return
    elif dress_type == "hair" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通头发"
        return
    elif dress_type == "hair" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图头发"
        return

    if dress_type == "shirt" and int(sprite_value[0]) == 1:
        ws.cell(2 + count, 4).value = "普通上衣"
        return
    elif dress_type == "shirt" and int(sprite_value[0]) > 1:
        ws.cell(2 + count, 4).value = "多贴图上衣"
        return





WriteAssetInfo()



#read prefab folder
#write prefab path



#read assets folder
#write assets paths
    #texture
    #mesh
    #material
    #sprite




