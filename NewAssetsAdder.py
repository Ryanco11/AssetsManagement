import openpyxl
import os
import shutil
import datetime
from pathlib import Path


##################################################################################################
##############################使用前请替换「unity工程」与「美术资源库」根目录路径##########################
##################################################################################################
#unity 工程项目根目录
avatar_root_path = r'/Users/ryanco/Projects/Master/wonder_party/avatarProject'
#art rescourse 根目录
art_root_path= r'/Users/ryanco/Projects/avatar_art_resources'
##################################################################################################
##################################################################################################
##################################################################################################



now = datetime.datetime.now()
year = '{:02d}'.format(now.year)
month = '{:02d}'.format(now.month)
day = '{:02d}'.format(now.day)
hour = '{:02d}'.format(now.hour)
minute = '{:02d}'.format(now.minute)
day_month_year = '{}-{}-{} {}:{}'.format(year, month, day, hour, minute)

print('day_month_year: ' + day_month_year)




excel_path = avatar_root_path + r'/Assets/Editor/AssetsManagement/SuitExcel/SuitExcel.xlsx'
project_path = avatar_root_path + r'/'
artsrc_path = art_root_path + r'/'

unity_prefab_path = r'Assets/Art/BundleResources/Dress/'
unity_asset_path = r'Assets/Art/model/coat/'
unity_sprite_path = r'Assets/Art/BundleResources/Sprites/'

asset_path = avatar_root_path + r'/Assets/Art/model/coat/'
prefab_path = avatar_root_path + r'/Assets/Art/BundleResources/Dress/'
asset_sock_path = avatar_root_path + r'/Assets/Art/BundleResources/Textures/'

fbx_path = art_root_path + r'/Dress/'
png_path = art_root_path + r'/Dress/'



#get excel
wb = openpyxl.load_workbook(excel_path)
ws = wb['New_Added_Assets_List']


def GetAssetList():
    fbx_list = []
    png_list = []

###fbx
    for r, d, f in os.walk(fbx_path):
        for file in f:
            # print(file)
            if file.lower().endswith(".fbx"):
                fbx_list.append(os.path.join(r, file).replace(project_path, ""))

### png
    for r, d, f in os.walk(png_path):
        for file in f:
            # print(file)
            if file.lower().endswith(".png"):
                png_list.append(os.path.join(r, file).replace(project_path, ""))

    return fbx_list, png_list

# def GetLastRow():
#     # last_row = 0
#     for row in range(2, ws.max_row):
#         print(ws.cell(row, 4).value)
#         if(ws.cell(row, 4).value is None):
#             last_row = row;
#             print("last_row: " + str(last_row))
#             break
#     return last_row

def GetLastRow():
    for row in range(1, ws.max_row + 100000):
        if (ws.cell(row, 4).value is None):
            last_row = row;  # this last_row value is actual plus one by actual last row in excel, cus py access col by minis one
            print("last row is : " + str(last_row))
            return last_row


def GetNewAssetInfo():

    #get all asset in art path
    # get source files [0]:fbx [1]:png
    sperate_asset_list = GetAssetList()
    fbx_list = sperate_asset_list[0]
    png_list = sperate_asset_list[1]

    asset_list = []

    for i in fbx_list:
        # print("fbx:" + i)
        asset_list.append(i)
    for i in png_list:
        # print("png:" + i)
        asset_list.append(i)

    for i in asset_list:
        print("all_assets: " + i)

    #get new added asset list
        #get last row from new excel
    last_row = GetLastRow()

    excel_namecode_list = []
    new_name_list = []
    mt_namecode_list = []
        #get namecode list from excel
    for row in range(2, last_row):
        excel_namecode_list.append(ws.cell(row, 4).value)
        print("ws: " + ws.cell(row, 4).value)

        #remove existing asset
    for asset in asset_list:
        for namecode in excel_namecode_list:
            if asset.__contains__(namecode):
                asset_list.remove(asset)
                break

    for i in asset_list:
        print("new_assets: " + i)
    new_asset_list = asset_list

    #sort by name code
        #get new name code list
    for asset in asset_list:
        mt_namecode_list.append(asset.split(r'/')[-2])
    print(mt_namecode_list)

        #remove mt name code
    for i in mt_namecode_list:
        if i not in new_name_list and i not in excel_namecode_list:
            new_name_list.append(i)

    print(new_name_list)

    return new_asset_list, new_name_list


def ProcessAssetInfo(new_asset_list ,new_name_list):
    last_row = GetLastRow()
    # print(last_row)
    # for each name code
    for namecode in new_name_list:
        print("t-nc:" + namecode)
        for asset in new_asset_list:
            print("t-asset:" + asset)
            if asset.__contains__(namecode):
                #for specfic file path with its namecode
                #write type
                if asset.__contains__(r'/A/') or asset.__contains__(r'/Q_长/'):
                    print("write in suit")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "suit"
                    ws.cell(last_row, 6).value = "普通套装"
                    png_text, fbx_text, sprite_text, muti = MoveFiles(namecode, new_asset_list, "suit")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图套装"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "suit/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/K/') or asset.__contains__(r'/Q_短/'):
                    print("write in pants")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "pants"
                    ws.cell(last_row, 6).value = "普通裤子"
                    png_text, fbx_text, sprite_text, muti = MoveFiles(namecode, new_asset_list, "pants")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图裤子"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "pants/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/T/'):
                    print("write in headwear")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "headwear"
                    ws.cell(last_row, 6).value = "普通头饰"
                    png_text, fbx_text,sprite_text,  muti = MoveFiles(namecode, new_asset_list, "headwear")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图头饰"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "headwear/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/F/'):
                    print("write in hair")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "hair"
                    ws.cell(last_row, 6).value = "普通头发"
                    png_text, fbx_text,sprite_text,  muti = MoveFiles(namecode, new_asset_list, "hair")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图头发"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "hair/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                # elif asset.__contains__(r'/M/'):
                #     print("write in hair")
                #     ws.cell(last_row, 4).value = namecode
                #     ws.cell(last_row, 5).value = "hair"
                #     ws.cell(last_row, 6).value = "帽子头发"
                #     png_text, fbx_text = MoveFiles(namecode, new_asset_list, "hair")
                #     ws.cell(last_row, 10).value = png_text
                #     ws.cell(last_row, 12).value = fbx_text
                #     mat_text = fbx_text.replace("FBX", "mat")
                #     mat_text = mat_text.replace("fbx", "mat")
                #     ws.cell(last_row, 14).value = mat_text
                #     ws.cell(last_row, 7).value = unity_prefab_path + "hair/" + namecode + ".prefab"
                #     last_row += 1
                #     break
                elif asset.__contains__(r'/QT/'):
                    print("write in bladic")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "baldric"
                    ws.cell(last_row, 6).value = "普通背包"
                    png_text, fbx_text, sprite_text, muti = MoveFiles(namecode, new_asset_list, "baldric")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图背包"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "baldric/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/W/'):
                    print("write in sock")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "socks"
                    ws.cell(last_row, 6).value = "普通袜子"
                    png_text, fbx_text, sprite_text = MovePng(namecode, new_asset_list, "socks")
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = "0"
                    ws.cell(last_row, 14).value = "0"
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "0"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/X/'):
                    print("write in shoes")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "shoes"
                    ws.cell(last_row, 6).value = "普通鞋子"
                    png_text, fbx_text,sprite_text,  muti = MoveFiles(namecode, new_asset_list, "shoes")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图鞋子"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "shoes/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/Y/'):
                    print("write in glasses")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "glasses"
                    ws.cell(last_row, 6).value = "普通眼镜"
                    png_text, fbx_text, sprite_text, muti = MoveFiles(namecode, new_asset_list, "glasses")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图眼镜"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "glasses/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break
                elif asset.__contains__(r'/S/'):
                    print("write in shirt")
                    ws.cell(last_row, 3).value = day_month_year
                    ws.cell(last_row, 4).value = namecode
                    ws.cell(last_row, 5).value = "shirt"
                    ws.cell(last_row, 6).value = "普通上衣"
                    png_text, fbx_text,sprite_text,  muti = MoveFiles(namecode, new_asset_list, "shirt")
                    if muti:
                        ws.cell(last_row, 6).value = "多贴图上衣"
                    ws.cell(last_row, 10).value = png_text
                    ws.cell(last_row, 12).value = fbx_text
                    mat_text = fbx_text.replace("FBX", "mat")
                    mat_text = mat_text.replace("fbx", "mat")
                    ws.cell(last_row, 14).value = mat_text
                    ws.cell(last_row, 16).value = sprite_text
                    ws.cell(last_row, 7).value = "1|-|" + unity_prefab_path + "shirt/" + namecode + ".prefab"
                    wb.save(excel_path)
                    last_row += 1
                    break


    wb.save(excel_path)

# def WriteSubType():

def MovePng(namecode, new_asset_list, dress_type):
    fbx_text = ""
    png_text = ""
    sprite_text = ""
    png_count = 0
    sprite_count = 0

    for asset in new_asset_list:
        if asset.__contains__(namecode) and asset.__contains__("_01"):
            muti = True

    for asset in new_asset_list:
        if asset.__contains__(namecode):
            ###normal asset
            path_to_create = asset_sock_path + dress_type

            # copy file to new folder
            shutil.copy2(asset, path_to_create)  # target filename is /dst/dir/file.ext

            if asset.lower().endswith(".png"):
                png_count += 1
                text = path_to_create + "/" + os.path.basename(asset)
                png_text += "|-|" + text.replace(project_path, "")

                if not asset.lower().__contains__("_mask") and not asset.__contains__("_H"):
                    sprite_count += 1
                    text = unity_sprite_path + dress_type + "/" + os.path.basename(asset)
                    sprite_text += "|-|" + text

    # fbx_text = str(fbx_count) + fbx_text
    png_text = str(png_count) + png_text
    sprite_text = str(sprite_count) + sprite_text
    print("sprite_text:" + sprite_text)
    return png_text, fbx_text, sprite_text



def MoveFiles(namecode, new_asset_list, dress_type):
    ### for all flies matchs namecode
    fbx_text = ""
    png_text = ""
    sprite_text = ""
    fbx_count = 0
    png_count = 0
    sprite_count = 0
    muti = False

    for asset in new_asset_list:
        if asset.__contains__(namecode) and asset.__contains__("_01"):
            muti = True

    for asset in new_asset_list:
        if asset.__contains__(namecode):
            #create folder in unity model folder

            if not muti:
                ###normal asset
                path_to_create = asset_path + dress_type + "/" + namecode
                if os.path.isdir(path_to_create):
                    print("Exists")
                else:
                    try:
                        os.mkdir(path_to_create)
                    except OSError:
                        print("Creation of the directory %s failed" % path_to_create)
                    else:
                        print("Successfully created the directory %s " % path_to_create)

                # copy file to new folder
                shutil.copy2(asset, path_to_create)  # target filename is /dst/dir/file.ext

            else:
                # mutli assets
                # for _01 _02... textures
                if asset.__contains__("_0") and dress_type != "hair" and asset.lower().endswith(".png"):
                    #create root folder
                    if not os.path.isdir(prefab_path + dress_type + "/" + namecode):
                        os.mkdir(prefab_path + dress_type + "/" + namecode)

                    #create texture fodler
                    pfb_path_to_create = prefab_path + dress_type + "/" + namecode + "/texture"
                    if os.path.isdir(pfb_path_to_create):
                        print("Exists")
                    else:
                        try:
                            os.mkdir(pfb_path_to_create)
                        except OSError:
                            print("Creation of the directory %s failed" % pfb_path_to_create)
                        else:
                            print("Successfully created the directory %s " % pfb_path_to_create)

                    # copy file to new folder
                    shutil.copy2(asset, pfb_path_to_create)  # target filename is /dst/dir/file.ext

                else:
                    ###normal asset
                    path_to_create = asset_path + dress_type + "/" + namecode
                    if os.path.isdir(path_to_create):
                        print("Exists")
                    else:
                        try:
                            os.mkdir(path_to_create)
                        except OSError:
                            print("Creation of the directory %s failed" % path_to_create)
                        else:
                            print("Successfully created the directory %s " % path_to_create)

                    # copy file to new folder
                    shutil.copy2(asset, path_to_create)  # target filename is /dst/dir/file.ext

            ##get png and fbx text
            if asset.lower().endswith(".fbx"):
                fbx_count += 1
                text = path_to_create + "/" + os.path.basename(asset)
                fbx_text += "|-|" + text.replace(project_path, "")
            elif asset.lower().endswith(".png"):
                png_count += 1
                if muti and asset.__contains__("_0"):
                    text = prefab_path + dress_type + "/" + namecode + "/texture" + "/" + os.path.basename(asset)
                    png_text += "|-|" + text.replace(project_path, "")
                else:
                    text = path_to_create + "/" + os.path.basename(asset)
                    png_text += "|-|" + text.replace(project_path, "")

                if not asset.lower().__contains__("_mask") and not asset.__contains__("_H"):
                    sprite_count += 1
                    text = unity_sprite_path + dress_type + "/" + os.path.basename(asset)
                    sprite_text += "|-|" + text

                # if not asset.__contains__("_0"):
                #     sprite_count -= 1
                #     sprite_text = sprite_text.replace(text, "")
                #     sprite_text = sprite_text.replace("|-|", "")

    fbx_text = str(fbx_count) + fbx_text
    png_text = str(png_count) + png_text
    sprite_text = str(sprite_count) + sprite_text
    print("sprite_text:" + sprite_text)
    return png_text, fbx_text, sprite_text, muti


# Start
new_asset_list, new_name_list = GetNewAssetInfo()
ProcessAssetInfo(new_asset_list, new_name_list)