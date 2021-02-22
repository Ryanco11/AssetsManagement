import openpyxl
import os

excel_path = r'/Users/ryanco/Desktop/资源元表/服饰元表Excel.xlsx'
project_art_path = r'/Users/ryanco/Projects/AndoidProject/wonder_party/avatarProject/Assets/Art'
project_path = r'/Users/ryanco/Projects/AndoidProject/wonder_party/avatarProject'

#get excel
wb = openpyxl.load_workbook(excel_path)
lws = wb['Lagecy_Assets_List']
nws = wb['New_Added_Assets_List']


def AccessAssetPath(last_row, ws):

    # walk through whole project
    i = 1
    for r, d, f in os.walk(project_art_path):
        for file in f:

            #get abs path
            file = os.path.join(r, file)

            #ingore .meta
            if file.lower().endswith(".meta"):
                continue

            print(str(i) + ": " + file)
            i += 1

            #每一个文件检查一遍ws
            for row in range(2, last_row):   # start at 2 , cus first row is not the actual info
                file_col_list = [7, 10, 12, 14, 16]
                for col in file_col_list:
                    # print("asset " + str(col) + ": " + ws.cell(row, col).value)
                    cell_value = ws.cell(row, col).value
                    ws.cell(row, col).value = CheckLost(cell_value, file)

    # walk through ws, check any lost
    for row in range(2, last_row):  # start at 2 , cus first row is not the actual info
        # namecode
        # print("this is asset: " +  ws.cell(row, 4).value)

        ### Assets
        file_col_list = [7, 10, 12, 14, 16]

        for col in file_col_list:
            if not ws.cell(row, col).value == "0":
                print("loas asset " + str(col) + ": " + ws.cell(row, col).value)


def CheckLost(cell_value, file):
    new_cell_value = ""

#跳过缺失文件
    if cell_value.__contains__("缺失"):
        return cell_value

#剔除存在的资源
    cell_path_list = cell_value.split('|-|')
    for path in cell_path_list:
        if len(path) < 5:
            # 绕过序号
            continue
        if file.__contains__(path):
            cell_path_list.remove(path)
            cell_path_list[0] = str(int(cell_path_list[0]) - 1)

#回填剩下的内容
    new_cell_value += cell_path_list[0]
    for path in cell_path_list:
        if len(path) < 5:
            # 绕过序号
            continue
        new_cell_value += "|-|"
        new_cell_value += path
    # print(new_cell_value)
    return str(new_cell_value)







def AccessAssetSpecificSetting(last_row, ws):
    for row in range(2, last_row):
        # namecode
        print("this is asset: " + ws.cell(row, 2).value)


def GetLastRow(ws):
    for row in range(1, ws.max_row + 100000):
        if (ws.cell(row, 4).value is None):
            last_row = row;  # this last_row value is actual plus one by actual last row in excel, cus py access col by minis one
            print("last row is : " + str(last_row))
            return last_row
# def CheckAssetLosing():







###Start
#1. Check Lagecy Assets
last_row = GetLastRow(lws)
AccessAssetPath(last_row, lws)

print("done")

#2. Cehck New Added Asset Since 2021-02-21









###########          Tips              ###########

#(row, col)

# loop row  # start at 2 , cus first row is not the actual info
# for row in range(2, 7):
#     for col in range(2, 3):  # (2, 3) only access to number 2
#         print(ws.cell(row, col).value)
